import numpy as np
import pandas as pd
import os
import matplotlib
from matplotlib import pyplot as plt
import sklearn
import openpyxl
import xlwt
import sklearn
from sklearn.datasets import make_classification
from sklearn.mixture import GaussianMixture


def potential_rate_paper_set(input_file, sheet_name, output_file, paper_num, set_num):
    """ 
    This function converts potential vs. capacity data to capacity vs c-rate data
    The Highest x-value (capacity [mAh]) from the charge/discharge graph.
    The input excel file has the format that each excel file has
    many sheets and each sheet, separated by charge/discharge C-rate,
    contains many papers. Each paper can have more than 1 set of capacity
    versus voltage data, each occupying 1 column (i.e. each set takes up two columns).
    The first row of a sheet specifies the paper code number, e.g. 'paper # 1', and
    the second row of a sheet spcifies the battery's set number, e.g. 'set # 1', and
    the third row of a sheet contains the quantity name and units, e.g. 'Capacity (mAh/g)'
    and 'Voltage (V)'.
    
    PARAMETERS
    ----------
    1) Excel file (file path) - string
    2) sheetnames - list
    3) Paper # - string
    4) Set # - integer: max number of sets
    RETURNS
    -------
    Capacity vs c-rate dataframe
    """
    # Determining which set # and the number of set lists from the excel file
    set_list = ['set #' + str(i) for i in range (1, set_num + 1)]
    # Test that the input 'set_num' is an integer
    assert type(set_num) == int, 'set_num must be an integer'
    # Dataframing the interested potential vs capacity excel sheet
    df = pd.read_excel(input_file, sheet_name, header = [0,1,2]) #
    # Merging multiple spreadsheets
    df_sheets = []
    for i in sheet_name:
        df_sheets.append(df[i])
    df_merged = pd.concat(df_sheets, axis = 1)
    # Get C-rate numerical values
    rates = []
    for name in sheet_name:
        rate = name.split("C_")[0]
        rates.append(rate)
    c_rate = pd.DataFrame({"C rate": rates})
    # Selecting maximum capacity values for each dataset and
    # concatnate with corresponding c-rates or current density
    caplist = []
    for i in set_list:
        set_i = (df_merged[paper_num, i])
        set_i_max = set_i["Capacity (mAh/g)"].max(axis=0).array
        caplist.append(c_rate)
        caplist.append(pd.DataFrame({"Capacity (mAh/g)": set_i_max}))
    df_cap_rate = pd.concat(caplist, axis = 1)
    # Test that the output is a dataframe
    assert type(c_rate) == type(df_cap_rate), 'The output must be a dataframe'
    # Exporting the converted dataframe to an excel file
    df_cap_rate.to_excel(output_file, sheet_name = paper_num, index=False, header=True)
    # Test that the sheet name is a string
    assert type(paper_num) == str, 'sheetname must be a string'
    print('saved succesfully to' + output_file)
    return df_cap_rate


def potential_rate_all(input_file, output_file):
    """ 
    This function converts and dataframes all voltage potential data and separates them
    by paper and set numbers such that the users can see what options they have.
    
    The input excel file has the format that each excel file has
    many sheets and each sheet, separated by charge/discharge C-rate,
    contains many papers. Each paper can have more than 1 set of capacity
    versus voltage data, each occupying 1 column (i.e. each set takes up two columns).
    The first row of a sheet specifies the paper code number, e.g. 'paper # 1', and
    the second row of a sheet spcifies the battery's set number, e.g. 'set # 1', and
    the third row of a sheet contains the quantity name and units, e.g. 'Capacity (mAh/g)'
    and 'Voltage (V)'.
    
    PARAMETERS
    ----------
    1) Excel file (file path) - string
    2) 
    RETURNS
    -------
    A voltage-capaity dataframe by paper and set number
    """
    # Read all sheets in the input file into a dictionary
    dict_excel = pd.read_excel(input_file, sheet_name=None, header = [0,1,2])
    sheetnames = list(dict_excel.keys())
    dict_excel[sheetnames[0]]
    df_cap_rate = pd.DataFrame()
    for i, sheetname in enumerate(sheetnames):
        df = dict_excel[sheetname]
        rate = sheetname.split("C_")[0]
        for headers, columnval in df.iteritems():
            paper_num, set_num, quan = headers
            # Takes only the capacity data
            if 'capacity' in quan or 'Capacity' in quan:
                max_cap = np.nanmax(columnval.values)
                # Must create the dataframe first before adding multiple headers
                newdf = pd.DataFrame([[rate, max_cap]])
                # Add headers
                newdf_header = [[paper_num, paper_num],[set_num, set_num],['C-rate', quan]]
                newdf.columns = newdf_header
                df_cap_rate = pd.concat([df_cap_rate, newdf])
    # apply lambda function to each column to drop Nans
    df_cap_rate_all = df_cap_rate.apply(lambda x: pd.Series(x.dropna().values))
    # Export dataframe to Excel
    df_cap_rate_all.to_excel(output_file, sheet_name = 'CaapacityRate', index=True, header=True)
    return df_cap_rate_all



def excel_merge(dataframe, xls_file, sheetname):
    """
    This function adds the converted dataframe to an existing excel file
    PARAMETERS
    ----------
    1) converted dataframe - dataframe
    2) Excel file (file path) - string
    3) sheetname - string
    RETURNS
    -------
    A new sheet in the excel file
    """
    # Exporting the converted dataframe to an excel file
    dataframe.to_excel(xls_file, sheet_name = sheetname, index=False, header=True)
    # Test that the sheet name is a string
    assert type(sheetname) == str, 'sheetname must be a string'
    print('saved succesfully to' + xls_file)
    return

def capacity_cycle(capacity_cycle_array, n, current_list, current_unit, capacity_unit):
    '''
    This function converts capacity-cycle data to capacity-rate plots.
    For different selections of paper/set, the user will have to adjust 'n_components' until the groups
    are properly grouped, as seen in the plot below
    Inputs
    - capacity_cycle_array: nx2 capacity vs cycle # array
    - n: number of C-rates, or 'stairs'
    - current_list: a list of current rates or current densities
    - current_unit: a text string of current unit
    - capacity_unit: a text string od capacity unit
    '''
    model = GaussianMixture(n_components = n)
    model.fit(capacity_cycle_array)
    ## Use the model to make predictions about which group each datapoint belongs to
    ## predictions stored as an np array with indexes corresponding to points, and values to their assigned class
    prediction = model.predict(capacity_cycle_array)
    ## np.array of the unique classes 
    clusters = np.unique(prediction)
    ## Plot the points now that they are grouped
    for cluster in clusters:
        row_ix = np.where(prediction == cluster)
        plt.scatter(capacity_cycle_array[row_ix, 0], capacity_cycle_array[row_ix, 1])
    plt.ylabel("Capacity "+capacity_unit, fontsize=16)
    plt.xlabel("Cycle #", fontsize=16)
    plt.show
    ## Return the means of each 'stair' and sort
    means = model.means_
    means = means[np.argsort(means[:, 0])]
    means_of_groups = means[:,1]
    means_of_groups = pd.DataFrame(means_of_groups)
    means_of_groups = means_of_groups.rename(columns={0: "Capacity"+capacity_unit})
    ## Get list of current rates or current densities
    current_list = pd.DataFrame(np.array(current_list))
    current_header = "Current " + current_unit
    current_list = current_list.rename(columns={0: current_header})
    capacity_vs_current_density_df = pd.concat([current_list, means_of_groups], axis = 1)
    return capacity_vs_current_density_df